import pandas as pd
import pandas_market_calendars as mcal
from openpyxl import load_workbook
from config import etfs_lst, etfs_hedge_date_dict, etfs_calendar_dict, hedge_info, start_date, end_date, excel_path
import datetime
pd.set_option("display.max_rows", 100)
pd.set_option("display.max_columns", 15)
pd.set_option("display.width", 1000)


def get_calendar(mkt_name, start_date, end_date):
    mkt_ex = mcal.get_calendar(mkt_name)
    mkt_ex_schedule = mkt_ex.schedule(start_date=start_date, end_date=end_date)
    mkt_trading_days = mkt_ex_schedule.index.date
    return mkt_trading_days.tolist()


def get_hedge_dict(holding_date, reminder, trading_days):
    t, d = reminder.split('__')
    d0, drift = d.split('+')
    holding_date = datetime.datetime.date(holding_date)
    idx = trading_days.index(holding_date)
    hedge_date = trading_days[idx + int(drift)]
    return hedge_date, t


def match_trading_days_from_reminders(holding_date, Fixing_reminder, FX_reminder, trading_days, hedge_info_i_t, holding_value):
    Fixing_hedge_dict = dict()
    FX_hedge_dict = dict()

    # 记录订单来源信息
    holding_date_str = str(holding_date.year) + '-' + str(holding_date.month) + '-' + str(holding_date.day)
    source = 'source:' + holding_date_str + '_' + str(holding_value)

    # hedge股指期货的提醒时间
    k1, v1 = get_hedge_dict(holding_date, Fixing_reminder, trading_days)
    # hedge股指期货的量和方向
    index_price = hedge_info_i_t.loc[holding_date, 'index_price']
    multiplier = hedge_info_i_t.loc[holding_date, 'multiplier']
    FX1 = hedge_info_i_t.loc[holding_date, 'FX1']
    pr = hedge_info_i_t.loc[holding_date, 'pr']
    etf_price = hedge_info_i_t.loc[holding_date, 'etf_price']
    etf_shares_per_future = index_price * multiplier * FX1 / pr / etf_price

    Fixing_hedge_num = - holding_value / etf_shares_per_future
    Fixing_hedge_num_str = 'Fixing_hedge_num:'+str(Fixing_hedge_num)

    Fixing_hedge_dict[k1] = 'handle_time:' + v1
    Fixing_hedge_dict[k1] += ';' + source
    Fixing_hedge_dict[k1] += ';' + Fixing_hedge_num_str

    # hedge货币对的量和方向
    FX_hedge_num = - holding_value * hedge_info_i_t.loc[holding_date, 'etf_price'] / hedge_info_i_t.loc[
        holding_date, 'FX2']

    # hedge货币对的提醒时间
    if ',' in FX_reminder:
        qdii_connect_lst = FX_reminder.split(',')

        FX_qdii_reminder = [i for i in qdii_connect_lst if 'QDII' in i][0]
        k3, v3 = get_hedge_dict(holding_date, FX_qdii_reminder, trading_days)
        percent_3 = float(v3.split('%')[0]) / 100
        FX_hedge_num_3_str = 'FX_hedge_num:' + str(FX_hedge_num * percent_3)
        FX_hedge_dict[k3] = 'handle_time:' + v3
        FX_hedge_dict[k3] += ';' + source
        FX_hedge_dict[k3] += ';' + FX_hedge_num_3_str

        FX_connect_reminder = [i for i in qdii_connect_lst if 'Connect' in i][0]
        k4, v4 = get_hedge_dict(holding_date, FX_connect_reminder, trading_days)
        percent_4 = float(v4.split('%')[0]) / 100
        FX_hedge_num_4_str = 'FX_hedge_num:' + str(FX_hedge_num * percent_4)
        FX_hedge_dict[k4] = 'handle_time:' + v4
        FX_hedge_dict[k4] += ';' + source
        FX_hedge_dict[k4] += ';' + FX_hedge_num_4_str
    else:
        k2, v2 = get_hedge_dict(holding_date, FX_reminder, trading_days)
        FX_hedge_num_str = 'FX_hedge_num:' + str(FX_hedge_num)
        FX_hedge_dict[k2] = 'handle_time:' + v2
        FX_hedge_dict[k2] += ';' + source
        FX_hedge_dict[k2] += ';' + FX_hedge_num_str

    return Fixing_hedge_dict, FX_hedge_dict


def get_etf_trading_days(etf_code, start_date, end_date, etfs_calendar_dict):
    opendate_mkt_lst = etfs_calendar_dict[str(etf_code)]
    if len(opendate_mkt_lst) == 1:
        trading_days = get_calendar(opendate_mkt_lst[0], start_date, end_date)
    if len(opendate_mkt_lst) > 1:
        trading_days = set()
        for i in opendate_mkt_lst:
            mkt_i_trading_days = get_calendar(i, start_date, end_date)
            trading_days = trading_days & set(mkt_i_trading_days) if len(trading_days) != 0 else set(mkt_i_trading_days)
        trading_days = sorted(list(trading_days))
    return trading_days


def update_hedge_info(date, etfs_lst, hedge_info, excel_path):
    pass


def match_Fixing_FX(etfs_lst, etfs_holding_dataframe, etfs_hedge_date_dataframe, start_date, end_date, etfs_calendar_dict, hedge_info, excel_path):
    etfs_trading_days = set()
    for col in etfs_lst:
        etf_i_trading_days = get_etf_trading_days(col, start_date, end_date, etfs_calendar_dict)
        etfs_trading_days = etfs_trading_days | set(etf_i_trading_days)
    etfs_trading_days = sorted(list(etfs_trading_days))

    Fixing_hedge_df = pd.DataFrame(index=etfs_trading_days, columns=etfs_lst).notna()*1
    FX_hedge_df = pd.DataFrame(index=etfs_trading_days, columns=etfs_lst).notna()*1
    for date in etfs_holding_dataframe.index:  # date: datetime.date(2023, 5, 15)
        """此处hedge信息仅提示使用"""
        update_hedge_info(date, etfs_lst, hedge_info, excel_path)  # 从数据源，截面获取并更新数据到指定sheet中
        hedge_info_df = pd.read_excel(excel_path, index_col=0, sheet_name=1)

        holding_temp = etfs_holding_dataframe.loc[date].dropna()
        for etf_code in holding_temp.index:  # etf_code: 513010 float
            holding_value = holding_temp.loc[etf_code]
            hedge_info_i_t = hedge_info_df[(hedge_info_df.index == date) & (hedge_info_df['etf_code'] == etf_code)]
            Fixing_reminder = etfs_hedge_date_dataframe.loc[str(etf_code), 'Fixing']
            FX_reminder = etfs_hedge_date_dataframe.loc[str(etf_code), 'FX']
            trading_days = get_etf_trading_days(etf_code, start_date, end_date, etfs_calendar_dict)
            Fixing_hedge_dict, FX_hedge_dict = match_trading_days_from_reminders(date, Fixing_reminder, FX_reminder, trading_days, hedge_info_i_t, holding_value)
            for k in Fixing_hedge_dict:
                if Fixing_hedge_df.loc[k, str(etf_code)] == 0:
                    Fixing_hedge_df.loc[k, str(etf_code)] = Fixing_hedge_dict[k]
                else:
                    Fixing_hedge_df.loc[k, str(etf_code)] += ',' + Fixing_hedge_dict[k]
            for k in FX_hedge_dict:
                if FX_hedge_df.loc[k, str(etf_code)] == 0:
                    FX_hedge_df.loc[k, str(etf_code)] = FX_hedge_dict[k]
                else:
                    FX_hedge_df.loc[k, str(etf_code)] += ',' + FX_hedge_dict[k]
    return Fixing_hedge_df, FX_hedge_df


def cal_repetitive_info(df):
    df_repetitive = df[[',' in x for x in df['info']]]
    df_no_repetitive = df[[',' not in x for x in df['info']]]
    info_repetitive_df = pd.DataFrame()
    if len(df_repetitive):
        for idx in df_repetitive.index:
            info_repetitive = df_repetitive.loc[idx, 'info']
            info_repetitive_lst = info_repetitive.split(',')
            for i, info in enumerate(info_repetitive_lst):
                info_repetitive_df.loc[i, 'date'] = df_repetitive.loc[idx, 'date']
                info_repetitive_df.loc[i, 'etf_code'] = df_repetitive.loc[idx, 'etf_code']
                info_repetitive_df.loc[i, 'info'] = info
    df_trans = pd.concat([info_repetitive_df, df_no_repetitive], ignore_index=True)
    return df_trans


def automation(etfs_lst, etfs_hedge_date_dict, excel_path, start_date, end_date, etfs_calendar_dict, hedge_info):
    etfs_holding_dataframe = pd.read_excel(excel_path, index_col=0, sheet_name=0)
    etfs_hedge_date_dataframe = pd.DataFrame(etfs_hedge_date_dict).T
    Fixing_hedge_df, FX_hedge_df = match_Fixing_FX(etfs_lst, etfs_holding_dataframe, etfs_hedge_date_dataframe, start_date, end_date, etfs_calendar_dict,
                                                   hedge_info, excel_path)

    Fixing_hedge_df = Fixing_hedge_df.stack().reset_index()
    Fixing_hedge_df.columns = ['date', 'etf_code', 'info']
    Fixing_hedge_df = Fixing_hedge_df[Fixing_hedge_df['info'] != 0].sort_values('date')

    FX_hedge_df = FX_hedge_df.stack().reset_index()
    FX_hedge_df.columns = ['date', 'etf_code', 'info']
    FX_hedge_df = FX_hedge_df[FX_hedge_df['info'] != 0].sort_values('date')

    Fixing_hedge_df_trans = cal_repetitive_info(Fixing_hedge_df)
    FX_hedge_df_trans = cal_repetitive_info(FX_hedge_df)

    wb = load_workbook(filename=excel_path)
    if 'Fixing_hedge_info' in wb.sheetnames:
        ws = wb['Fixing_hedge_info']
        wb.remove(ws)
        wb.save(excel_path)

    wb = load_workbook(filename=excel_path)
    if 'FX_hedge_info' in wb.sheetnames:
        ws = wb['FX_hedge_info']
        wb.remove(ws)
        wb.save(excel_path)

    with pd.ExcelWriter(excel_path, mode='a') as writer:
        Fixing_hedge_df_trans.to_excel(writer, sheet_name='Fixing_hedge_info')
        FX_hedge_df_trans.to_excel(writer, sheet_name='FX_hedge_info')


if __name__ == '__main__':
    automation(etfs_lst, etfs_hedge_date_dict, excel_path, start_date, end_date, etfs_calendar_dict, hedge_info)